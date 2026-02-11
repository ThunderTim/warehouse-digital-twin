#!/usr/bin/env python3
"""


Generate mock inventory JSON from an Excel sheet.

- Reads Column G up to row 701 (inclusive) for bin/location strings
- Normalizes to 6-char bin IDs:
    * If trailing letter exists (e.g. 3E01A1A) -> drop it => 3E01A1
    * If ends with two digits and first is '0' (e.g. 3W34A03) -> compress => 3W34A3
- Dedupes bins (preserving order)
- Generates N mock inventory items with those bins
"""

from __future__ import annotations
import argparse
import json
import random
import re
import string
from typing import List, Optional

from openpyxl import load_workbook


# -----------------------------
# Bin normalization
# -----------------------------
TRAILING_LETTER_RE = re.compile(r"^([0-9][EW].*[0-9])([A-Z])$")  # ends with a letter after a digit
ENDS_TWO_DIGITS_RE = re.compile(r"^(.+?)(\d{2})$")               # ends with two digits


def normalize_bin(raw: str) -> Optional[str]:
    """
    Normalize a raw bin string to a 6-char identifier where possible.
    Returns None if it doesn't look like a bin code.
    """
    s = (raw or "").strip().upper()
    if not s:
        return None

    # skip obvious header text
    if s in {"STORAGE BIN", "BIN", "LOCATION"}:
        return None

    # 1) Drop trailing letter (ex: 3E01A1A -> 3E01A1)
    m = TRAILING_LETTER_RE.match(s)
    if m:
        s = m.group(1)

    # 2) If ends with two digits and leading is 0, compress (ex: 3W34A03 -> 3W34A3)
    m2 = ENDS_TWO_DIGITS_RE.match(s)
    if m2:
        prefix, dd = m2.group(1), m2.group(2)
        if dd.startswith("0"):
            s = prefix + dd[1:]  # drop leading zero

    # If still longer than 6, we won't hard-truncate (could corrupt identifiers).
    # But if it's exactly 6, great.
    # If it's 7+ and you truly want ONLY 6, enforce a stricter rule here.
    return s


# -----------------------------
# Random generators
# -----------------------------
ALPHANUM = string.ascii_letters + string.digits


def rand_base62(n: int) -> str:
    return "".join(random.choice(ALPHANUM) for _ in range(n))


def rand_doc() -> str:
    return f"{rand_base62(8)}-{random.randint(100000, 999999)}"


def rand_skpart() -> str:
    return "P" + f"{random.randint(0, 9999999):07d}"


def rand_innumb() -> str:
    return "I" + f"{random.randint(0, 9999999):07d}"


def rand_dim(minv: int, maxv: int) -> str:
    return str(random.randint(minv, maxv))


# -----------------------------
# Main
# -----------------------------
def read_bins_from_excel(
    excel_path: str,
    sheet_name: Optional[str],
    col_letter: str,
    max_row: int,
) -> List[str]:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    col = ord(col_letter.upper()) - ord("A") + 1  # 'G' -> 7

    bins: List[str] = []
    for r in range(1, max_row + 1):
        v = ws.cell(row=r, column=col).value
        if v is None:
            continue
        norm = normalize_bin(str(v))
        if not norm:
            continue
        bins.append(norm)

    # dedupe preserving order
    seen = set()
    uniq = []
    for b in bins:
        if b not in seen:
            seen.add(b)
            uniq.append(b)

    return uniq


def make_inventory_items(bins: List[str], n: int, seed: int) -> List[dict]:
    random.seed(seed)

    if not bins:
        raise RuntimeError("No bins found after parsing/normalizing.")

    # If you ask for more items than bins, we cycle through bins
    items: List[dict] = []
    for i in range(n):
        bin_id = bins[i % len(bins)]
        items.append(
            {
                "indocn": rand_doc(),
                "lofull": str(random.choice([0, 25, 50, 75, 100])),
                "inavlq": str(random.randint(0, 200)),
                "itemdp": rand_dim(6, 60),
                "itemht": rand_dim(6, 72),
                "itemwd": rand_dim(6, 60),
                "lolocn": bin_id,
                "skskun": rand_base62(16),
                "skpart": rand_skpart(),
                "innumb": rand_innumb(),
                "imageUrl": "",
                "status": "success",
            }
        )
    return items


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", required=True, help="Path to the Excel file")
    ap.add_argument("--sheet", default=None, help="Sheet name (default: active)")
    ap.add_argument("--col", default="G", help="Column letter for bins (default: G)")
    ap.add_argument("--max-row", type=int, default=701, help="Max row to read (default: 701)")
    ap.add_argument("--n", type=int, default=100, help="How many inventory items to generate")
    ap.add_argument("--seed", type=int, default=22, help="Random seed for repeatability")
    ap.add_argument("--out", default="inventory_mock.json", help="Output JSON path")
    args = ap.parse_args()

    bins = read_bins_from_excel(args.excel, args.sheet, args.col, args.max_row)

    # OPTIONAL: enforce ONLY length-6 bin IDs (uncomment if you want strict enforcement)
    # bins = [b for b in bins if len(b) == 6]

    items = make_inventory_items(bins, args.n, args.seed)

    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(items, f, indent=2)

    print(f"✅ Wrote {len(items)} items to {args.out}")
    print(f"📍 Unique bins parsed: {len(bins)}")
    print("🔎 First 10 bins:", bins[:10])


if __name__ == "__main__":
    main()
